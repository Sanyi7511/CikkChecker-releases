"""
CikkChecker v3 – pywebview + Rust backend
Modern gradient web UI with native window.
"""
import json, os, queue, subprocess, sys, tempfile, threading, time
from collections import Counter
from datetime import datetime
from tkinter import filedialog
import tkinter as tk

import requests
import webview

# Version is read from version.txt (written by GitHub Actions at build time)
def _read_version():
    try:
        vpath = os.path.join(os.path.dirname(os.path.abspath(__file__)), "version.txt")
        with open(vpath, encoding="utf-8") as f:
            return f.read().strip().lstrip("v")
    except:
        return "0.0.0"

APP_VERSION = _read_version()
UPDATE_REPO  = "Sanyi7511/CikkChecker-releases"
UPDATE_ASSET = "CikkCheckerSetup.exe"
BINARY_NAME  = "checker_core.exe" if sys.platform == "win32" else "checker_core"
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
BINARY_PATH  = os.path.join(BASE_DIR, BINARY_NAME)
CONFIG_PATH  = os.path.join(BASE_DIR, "config.json")
UI_PATH      = os.path.join(BASE_DIR, "ui", "index.html")

DEFAULT_CONFIG = {
    "theme": "dark", "language": "hu",
    "auto_update": True, "update_interval": "on_start",
    "sleep_seconds": 0.8, "save_every": 100,
    "sort_order": "abc", "show_price": True,
    "auto_rm_dup": False, "dont_ask_dup": False, "save_dup_txt": True,
    "last_url": "", "last_xlsx": "", "last_txt": "",
}

def load_config():
    try:
        if os.path.exists(CONFIG_PATH):
            with open(CONFIG_PATH, encoding="utf-8") as f:
                return {**DEFAULT_CONFIG, **json.load(f)}
    except: pass
    return dict(DEFAULT_CONFIG)

def save_config_file(cfg):
    try:
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=2, ensure_ascii=False)
    except: pass


def export_excel(csv_path, excel_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        wb = Workbook(); ws = wb.active; ws.title = "CikkChecker"
        fills = {
            "Van":          PatternFill("solid", fgColor="0A2016"),
            "Nincs":        PatternFill("solid", fgColor="200A0A"),
            "Külső raktár": PatternFill("solid", fgColor="201508"),
        }
        fonts = {
            "Van":          Font(color="23C55E", bold=True, size=10),
            "Nincs":        Font(color="F04747", bold=True, size=10),
            "Külső raktár": Font(color="F5A623", bold=True, size=10),
        }
        thin = Side(style="thin", color="1F2535")
        brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.append(["Cikkszám","Elérhetőség","Ár"])
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="0D1018")
            cell.font = Font(color="9BA3B4", bold=True, size=10)
            cell.alignment = Alignment(horizontal="center"); cell.border = brd
        import csv as _csv
        if os.path.exists(csv_path):
            with open(csv_path, encoding="utf-8-sig", newline="") as f:
                for row in _csv.DictReader(f):
                    # Try both accented and plain column names
                    c = (row.get("Cikkszám") or row.get("Cikkszam") or "").strip()
                    a = (row.get("Elérhetőség") or row.get("Elerhetoseg") or "").strip()
                    p = (row.get("Ár") or row.get("Ar") or "").strip()
                    if not c: continue
                    ws.append([c, a, p]); r = ws.max_row
                    for col in range(1, 4):
                        cell = ws.cell(r, col); cell.border = brd
                        cell.alignment = Alignment(horizontal="left" if col==1 else "center")
                        cell.fill = fills.get(a, PatternFill("solid", fgColor="13161F"))
                    ws.cell(r,2).font = fonts.get(a, Font(color="9BA3B4", size=10))
                    ws.cell(r,1).font = Font(color="E8EBF0", size=10)
                    ws.cell(r,3).font = Font(color="9BA3B4", size=10)
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 16
        wb.save(excel_path)
    except Exception as e:
        print(f"Excel export error: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# Python API — exposed to JavaScript via pywebview
# ══════════════════════════════════════════════════════════════════════════════
class Api:
    def __init__(self):
        self._window     = None   # set after window creation
        self._cfg        = load_config()
        self._proc       = None
        self._stop_req   = False
        self._unk_mode   = None   # None | "skip_all" | "stop_all"
        self._unk_event  = None
        self._unk_result = None
        self._xlsx_path  = ""
        self._csv_path   = ""

    def set_window(self, w):
        self._window = w

    # ── Config ────────────────────────────────────────────────────────────
    def get_config(self):
        cfg = dict(self._cfg)
        cfg["app_version"] = APP_VERSION
        cfg["scheduler"]   = self.get_scheduler()
        return cfg

    def save_config(self, cfg):
        self._cfg.update(cfg)
        save_config_file(self._cfg)
        return True

    # ── File dialogs (must run on main thread via tk) ─────────────────────
    def _tk_dialog(self, fn):
        result = [None]
        ev = threading.Event()
        def _run():
            root = tk.Tk()
            root.withdraw()
            try: root.attributes('-topmost', True)
            except: pass
            result[0] = fn(root)
            root.destroy(); ev.set()
        threading.Thread(target=_run, daemon=True).start()
        ev.wait(timeout=30)
        return result[0]

    def browse_file(self, filetype: str):
        """Open file dialog for txt/csv/xlsx."""
        type_map = {
            "txt":  [("Text fájl", "*.txt"), ("Minden", "*.*")],
            "csv":  [("CSV fájl", "*.csv"), ("Minden", "*.*")],
            "xlsx": [("Excel fájl", "*.xlsx *.xls"), ("Minden", "*.*")],
        }
        filetypes = type_map.get(filetype, [("Minden", "*.*")])
        path = self._tk_dialog(lambda r: filedialog.askopenfilename(
            parent=r,
            title=f"Cikkszám forrás megnyitása ({filetype.upper()})",
            filetypes=filetypes))
        if path:
            self._cfg["last_txt"] = path
            self._cfg["last_txt_type"] = filetype
            save_config_file(self._cfg)
        return path or ""

    def parse_file(self, path: str, filetype: str, col_index: int = 0):
        """Parse codes from txt/csv/xlsx. Returns {codes, preview} or None."""
        try:
            codes = []
            preview = []

            if filetype == "txt":
                with open(path, encoding="utf-8", errors="replace") as f:
                    codes = [l.strip() for l in f if l.strip()]

            elif filetype == "csv":
                import csv as _csv
                with open(path, encoding="utf-8-sig", errors="replace") as f:
                    sample = f.read(2048)
                # Try to detect delimiter, fallback to semicolon then comma
                try:
                    dialect = _csv.Sniffer().sniff(sample, delimiters=",;	|")
                except _csv.Error:
                    dialect = _csv.excel  # default comma
                with open(path, encoding="utf-8-sig", errors="replace", newline="") as f:
                    reader = _csv.reader(f, dialect)
                    rows = list(reader)
                # Skip header if first cell looks like a header (non-numeric text)
                start = 0
                if rows and not any(c.isdigit() for c in (rows[0][col_index] if col_index < len(rows[0]) else "")):
                    start = 1
                codes   = [row[col_index].strip() for row in rows[start:] if col_index < len(row) and row[col_index].strip()]
                preview = [row[col_index].strip() for row in rows[start:start+5] if col_index < len(row)]

            elif filetype in ("xlsx", "xls"):
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(path, read_only=True, data_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    wb.close()
                    # Skip header if first cell is text (not a code)
                    start = 0
                    if rows:
                        first = str(rows[0][col_index] or "").strip()
                        if first and not any(c.isdigit() for c in first):
                            start = 1
                    codes   = [str(row[col_index]).strip() for row in rows[start:]
                               if col_index < len(row) and row[col_index] is not None
                               and str(row[col_index]).strip()]
                    preview = [str(row[col_index]).strip() for row in rows[start:start+5]
                               if col_index < len(row) and row[col_index] is not None]
                except ImportError:
                    return {"error": "openpyxl nincs telepítve"}

            # Deduplicate while preserving order
            seen_set = set()
            unique = [c for c in codes if not (c in seen_set or seen_set.add(c))]

            return {"codes": unique, "preview": preview, "total": len(codes), "unique": len(unique)}

        except Exception as e:
            self._js(f"onLog({json.dumps(f'[ERR] Fájl parse hiba: {e}')})")
            return None

    def drop_file(self, filename: str, filetype: str):
        """pywebview can't access drag-drop file paths directly — return None,
        user should use the browse buttons instead."""
        return None

    def browse_txt(self):
        """Legacy compat — redirect to browse_file."""
        return self.browse_file("txt")

    def export_excel_dialog(self, data):
        """Called from JS with table row data — shows save dialog and exports Excel."""
        path = self._tk_dialog(lambda r: filedialog.asksaveasfilename(
            parent=r, title="Excel mentési hely",
            initialfile=f"CikkChecker_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            defaultextension=".xlsx", filetypes=[("Excel fájl", "*.xlsx")]))
        if not path:
            return None
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            wb = Workbook(); ws = wb.active; ws.title = "CikkChecker"
            fills = {
                "Van":          PatternFill("solid", fgColor="0A2016"),
                "Nincs":        PatternFill("solid", fgColor="200A0A"),
                "Külső raktár": PatternFill("solid", fgColor="201508"),
            }
            fonts = {
                "Van":          Font(color="23C55E", bold=True, size=10),
                "Nincs":        Font(color="F04747", bold=True, size=10),
                "Külső raktár": Font(color="F5A623", bold=True, size=10),
            }
            thin = Side(style="thin", color="2A3147")
            brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws.append(["Cikkszám", "Elérhetőség", "Ár"])
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor="0D1018")
                cell.font = Font(color="9BA3B4", bold=True, size=10)
                cell.alignment = Alignment(horizontal="center"); cell.border = brd
            for row in (data or []):
                c = str(row.get("code","")).strip()
                a = str(row.get("avail","")).strip()
                p = str(row.get("price","")).strip()
                if not c: continue
                ws.append([c, a, p]); r = ws.max_row
                for col in range(1,4):
                    cell = ws.cell(r, col); cell.border = brd
                    cell.alignment = Alignment(horizontal="left" if col==1 else "center")
                    cell.fill = fills.get(a, PatternFill("solid", fgColor="13161F"))
                ws.cell(r,2).font = fonts.get(a, Font(color="9BA3B4", size=10))
                ws.cell(r,1).font = Font(color="E8EBF0", size=10)
                ws.cell(r,3).font = Font(color="9BA3B4", size=10)
            ws.column_dimensions["A"].width = 28
            ws.column_dimensions["B"].width = 18
            ws.column_dimensions["C"].width = 16
            wb.save(path)
            return path
        except Exception as e:
            self._js(f"onLog({json.dumps(f'[ERR] Excel export hiba: {e}')})")
            return None

    # ── Login detection ───────────────────────────────────────────────────
    def detect_login(self, url):
        try:
            if not url.startswith("http"): url = "https://" + url
            resp = requests.get(f"{url}/product-search/test/0?1", timeout=8, allow_redirects=True)
            html = resp.text.lower()
            return ("/login" in resp.url or
                    any(k in html for k in ["bejelentkezés","belépés","felhasználónév","user-name","password"]))
        except:
            return False

    # ── Process control ───────────────────────────────────────────────────
    def start_process(self, config):
        if self._proc and self._proc.poll() is None:
            self._js("onLog('[!!] A folyamat már fut.')")
            return False

        url        = config.get("url","").strip()
        user       = config.get("username","").strip()
        pw         = config.get("password","").strip()
        req_login  = config.get("req_login", False)
        txt_path   = config.get("txt_path","").strip()
        sleep_s    = float(config.get("sleep_s", 0.8))
        save_every = int(config.get("save_every", 100))
        sort_order = config.get("sort_order","abc")
        auto_rm    = config.get("auto_rm", False)
        save_dup   = config.get("save_dup", True)
        manual     = config.get("manual","").strip()
        skip_dup   = config.get("skip_dup", False)

        if not url:
            self._js("onLog('[!!] Add meg az oldal URL-jét.')")
            return False
        if req_login and (not user or not pw):
            self._js("onLog('[!!] Felhasználónév és jelszó szükséges.')")
            return False

        # Auto-determine output folder:
        # 1. Same folder as input file (if given)
        # 2. Desktop
        # 3. Current directory
        if txt_path and os.path.exists(txt_path):
            out_folder = os.path.dirname(os.path.abspath(txt_path))
        else:
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            out_folder = desktop if os.path.exists(desktop) else os.path.abspath(".")

        ts = datetime.now().strftime("%Y%m%d_%H%M")
        xlsx_path = os.path.join(out_folder, f"CikkChecker_{ts}.xlsx")

        # Collect codes — preloaded takes priority over raw txt_path
        preloaded = config.get("preloaded", [])
        codes = []
        if preloaded:
            codes = [str(c).strip() for c in preloaded if str(c).strip()]
        elif txt_path and os.path.exists(txt_path):
            # Fallback: read raw txt
            with open(txt_path, encoding="utf-8", errors="replace") as f:
                codes += [l.strip() for l in f if l.strip()]
        if manual:
            codes += [l.strip() for l in manual.splitlines() if l.strip()]
        if not codes:
            self._js("onLog('[!!] Adj meg cikkszámokat.')")
            return False

        xlsx_folder = os.path.dirname(xlsx_path) or "."
        os.makedirs(xlsx_folder, exist_ok=True)
        csv_path = os.path.join(xlsx_folder, "cikkchecker_results.csv")

        # Duplicates
        counter   = Counter(codes)
        dup_items = {c:n for c,n in counter.items() if n>1}
        unique    = list(dict.fromkeys(codes))

        if dup_items and not skip_dup:
            dup_rows = sum(n-1 for n in counter.values() if n>1)
            if auto_rm:
                codes = unique
                self._js(f"onLog('[dup] {dup_rows} duplikált sor eltávolítva.')")
            if save_dup:
                dup_path = os.path.join(xlsx_folder, "duplicates.txt")
                with open(dup_path,"w",encoding="utf-8") as f:
                    for c,n in dup_items.items(): f.write(f"{c};{n}\n")

        # Total
        self._js(f"document.getElementById('s-total').textContent = '{len(set(codes))}'")

        # Write codes to temp file
        tmp = tempfile.NamedTemporaryFile(mode="w", suffix=".txt",
                                          encoding="utf-8", delete=False)
        tmp.write("\n".join(codes)); tmp.close()

        self._cfg["last_url"] = url
        self._cfg["last_txt"] = txt_path
        save_config_file(self._cfg)

        if not os.path.exists(BINARY_PATH):
            self._js(f"onLog('[!!] checker_core nem található: {BINARY_PATH}')")
            return False

        cmd = [BINARY_PATH,
               "--base-url", url,
               "--user", user, "--password", pw,
               "--codes-file", tmp.name, "--csv-output", csv_path,
               "--sleep-seconds", str(sleep_s), "--save-every", str(save_every),
               "--sort-order", sort_order]
        if req_login: cmd.append("--requires-login")

        self._stop_req   = False
        self._unk_mode   = None
        self._xlsx_path  = xlsx_path
        self._csv_path   = csv_path

        self._js(f"onLog('[--] Indítás — {len(codes)} cikkszám')")
        self._js(f"onLog('[--] Kimenet: {xlsx_path}')")
        self._js("onStatus('Indítás...')")

        self._proc = subprocess.Popen(
            cmd, stdin=subprocess.PIPE, stdout=subprocess.PIPE,
            stderr=subprocess.PIPE, text=True, bufsize=1)

        threading.Thread(target=self._read_proc, args=(tmp.name,), daemon=True).start()
        return True

    def stop_process(self):
        self._stop_req = True
        if self._proc:
            try: self._proc.stdin.write("stop\n"); self._proc.stdin.flush()
            except: pass
        self._js("onLog('[--] Leállítás kérve...')")
        return True

    def restart_process(self):
        self.stop_process()
        self._stop_req  = False
        self._unk_mode  = None
        self._js("onLog('[--] Újraindítás — folyamat leállítva, újraindítható.')")
        self._js("onStatus('Várakozás')")
        return True

    # ── Process reader ────────────────────────────────────────────────────
    def _read_proc(self, tmp_path):
        try:
            for line in self._proc.stdout:
                line = line.strip()
                if not line: continue
                try: msg = json.loads(line)
                except: self._js(f"onLog({json.dumps(line)})"); continue

                kind = msg.get("kind","")

                if kind == "log":
                    ts = datetime.now().strftime("%H:%M:%S")
                    text = msg.get("msg", "")
                    self._js(f"onLog({json.dumps(f'[{ts}] {text}')})")

                elif kind == "status":
                    self._js(f"onStatus({json.dumps(msg['msg'])})")

                elif kind == "progress":
                    self._js(f"onProgress({msg['current']},{msg['total']})")

                elif kind == "login_detected":
                    req = str(msg.get("required",False)).lower()
                    self._js(f"onLoginDetected({req})")

                elif kind == "result":
                    code  = json.dumps(msg.get("cikkszam",""))
                    avail = json.dumps(msg.get("elerhetoseg",""))
                    price = json.dumps(msg.get("ar",""))
                    self._js(f"onResult({code},{avail},{price})")

                elif kind == "unknown":
                    code = msg["cikkszam"]
                    action = self._handle_unknown(code)
                    try: self._proc.stdin.write(action+"\n"); self._proc.stdin.flush()
                    except: pass

                elif kind == "error":
                    self._js(f"onLog({json.dumps('[ERR] ' + msg['msg'])})")

                elif kind == "done":
                    self._js("onLog('[--] Kész — Excel exportálás...')")
                    export_excel(self._csv_path, self._xlsx_path)
                    self._js(f"onLog({json.dumps(f'[--] Excel mentve: {self._xlsx_path}')})")
                    self._js("onDone()")

        except Exception as e:
            self._js(f"onLog({json.dumps(f'[ERR] Olvasási hiba: {e}')})")
        finally:
            try: os.unlink(tmp_path)
            except: pass

    def _handle_unknown(self, code):
        if self._unk_mode == "skip_all": return "skip"
        if self._unk_mode == "stop_all": return "stop"

        self._unk_event  = threading.Event()
        self._unk_result = None
        self._js(f"onUnknown({json.dumps(code)}).then(r => {{ window._unkResult = r; }})")

        # Poll for JS result
        for _ in range(300):  # 30 seconds max
            time.sleep(0.1)
            if self._unk_result is not None:
                break

        result = self._unk_result or {"action":"stop","applyAll":False}
        if result.get("applyAll"):
            if result["action"] == "skip": self._unk_mode = "skip_all"
            else:                          self._unk_mode = "stop_all"
        return result["action"]

    def set_unknown_result(self, action, apply_all):
        """Called from JS after user resolves unknown dialog."""
        self._unk_result = {"action": action, "applyAll": apply_all}
        return True

    # ── Update ────────────────────────────────────────────────────────────
    def check_update(self):
        threading.Thread(target=self._do_check_update, daemon=True).start()
        return True

    def _do_check_update(self):
        try:
            r = requests.get(
                f"https://api.github.com/repos/{UPDATE_REPO}/releases/latest",
                timeout=10)
            r.raise_for_status(); data = r.json()
        except:
            self._js("onLog('[--] Frissítés ellenőrzése sikertelen.')")
            return

        latest = data.get("tag_name","").strip().lstrip("v")
        notes  = data.get("body","").strip()

        def _v(s):
            s = s.strip().lstrip("v")
            parts = s.split(".")
            return tuple(int(x) if x.isdigit() else 0 for x in parts)

        current = _v(APP_VERSION)
        remote  = _v(latest)

        self._js(f"onLog('[--] Jelenlegi verzió: {APP_VERSION} | GitHub: {latest}')")

        if not latest or remote <= current:
            self._js("onLog('[--] Nincs újabb verzió.')")
            return

        # Find installer asset — name includes version e.g. CikkCheckerSetup_3.1.0.exe
        asset_url = next((a["browser_download_url"] for a in data.get("assets",[])
                          if a.get("name","").startswith("CikkCheckerSetup")), None)
        if not asset_url:
            self._js("onLog('[--] Új verzió van, de a telepítő nem található.')")
            return

        self._js(f"onLog('[!!] Új verzió elérhető: v{latest}')")
        # Show banner — wait for user to approve before downloading
        self._js(f"onUpdateAvailable({json.dumps(latest)},{json.dumps(asset_url)})")

    def do_update(self, asset_url: str, latest: str):
        """Called from JS when user clicks the update button."""
        threading.Thread(target=self._dl_update,
                         args=(asset_url, latest), daemon=True).start()
        return True

    def _dl_update(self, url, latest):
        self._js(f"onLog('[--] Frissítő letöltése: v{latest}')")
        self._js("onStatus('Frissítés letöltése...')")
        self._js(f"onUpdateDownloading({json.dumps(latest)})")
        try:
            tmp_dir  = tempfile.gettempdir()
            # Versioned filename to avoid conflicts with running instance
            versioned_name = f"CikkCheckerSetup_{latest}.exe"
            tmp = os.path.join(tmp_dir, versioned_name)

            with requests.get(url, stream=True, timeout=120) as r:
                r.raise_for_status()
                total = int(r.headers.get("content-length", 0))
                downloaded = 0
                with open(tmp, "wb") as f:
                    for chunk in r.iter_content(8192):
                        if chunk:
                            f.write(chunk)
                            downloaded += len(chunk)
                            if total:
                                pct = int(downloaded / total * 100)
                                self._js(f"onUpdateProgress({pct})")

            self._js("onLog('[--] Letöltés kész — telepítő indul...')")
            self._js("onUpdateReady()")
            time.sleep(1.5)

            # Use a .bat launcher with delay so the app can fully close
            # before the installer touches any DLL files
            bat_path = os.path.join(tmp_dir, "cikkchecker_update.bat")
            # Find the installed CikkChecker.exe path to restart after update
            app_exe = os.path.abspath(sys.executable if getattr(sys, "frozen", False) else __file__)
            # If running as PyInstaller bundle, exe is sys.executable
            # If running as script, find the CikkChecker.exe in same dir
            if not getattr(sys, "frozen", False):
                app_exe = os.path.join(os.path.dirname(os.path.abspath(__file__)), "CikkChecker.exe")

            bat = (
                "@echo off\n"
                ":: Varunk amig az app teljesen bezarul\n"
                "timeout /t 4 /nobreak > nul\n"
                ":: Csendes telepites\n"
                f'start "" /wait "{tmp}" /VERYSILENT /SUPPRESSMSGBOXES /NORESTART /SP- /CLOSEAPPLICATIONS\n'
                ":: Varunk hogy a telepito befejezze\n"
                "timeout /t 3 /nobreak > nul\n"
                ":: Ujrainditjuk az appot\n"
                f'if exist "{app_exe}" start "" "{app_exe}"\n'
                ":: Self-delete\n"
                "del \"%~f0\"\n"
            )
            with open(bat_path, "w", encoding="utf-8") as f:
                f.write(bat)

            import subprocess as sp
            sp.Popen(
                ["cmd", "/c", bat_path],
                creationflags=0x08000000,  # CREATE_NO_WINDOW
                close_fds=True
            )
            time.sleep(0.5)
            self._window.destroy()

        except Exception as e:
            self._js(f"onLog({json.dumps(f'[ERR] Frissítési hiba: {e}')})")
            self._js("onUpdateFailed()")

    # ── Kód kinyerő (Code Extractor) ─────────────────────────────────────
    def scrape_codes(self, urls: list, selector: str, max_pages: int,
                     user: str = "", password: str = "") -> dict:
        """
        Scrape product codes from source website(s).
        Strategies (tried in order):
          1. Excel/CSV export button detection → download & parse
          2. Specific CSS selector (if provided)
          3. Auto-detect common SKU HTML patterns
          4. Regex fallback on page text
        """
        """Scrape product codes from source website(s)."""
        import re as _re, io as _io
        try:
            from bs4 import BeautifulSoup as BS
        except ImportError:
            return {"error": "beautifulsoup4 nincs telepítve", "codes": [], "pages": 0}

        all_codes = []
        total_pages = 0

        CODE_PATTERN = _re.compile(r'\b([A-Z][A-Z0-9]{1,8}[-_./]?[A-Z0-9]{2,15}|[0-9]{5,15})\b')

        def extract_from_jsonld(html: str) -> list:
            """Extract SKU/MPN/productID from JSON-LD schema.org scripts."""
            import json as _json
            codes = []
            try:
                soup_jl = BS(html, "html.parser")
                for script in soup_jl.find_all("script", type="application/ld+json"):
                    try:
                        data = _json.loads(script.string or "")
                        items = data if isinstance(data, list) else [data]
                        for item in items:
                            # Handle @graph
                            if "@graph" in item:
                                items.extend(item["@graph"])
                            t = item.get("@type","")
                            if t in ("Product","ItemList","ListItem") or "product" in str(t).lower():
                                for field in ["sku","mpn","productID","gtin13","gtin8","identifier","code"]:
                                    val = item.get(field,"")
                                    if val and 2 < len(str(val)) <= 40:
                                        codes.append(str(val).strip())
                                # Check offers
                                offers = item.get("offers", [])
                                if isinstance(offers, dict): offers = [offers]
                                for o in offers:
                                    for f in ["sku","mpn","itemOffered"]:
                                        val = o.get(f,"")
                                        if val and isinstance(val,str) and 2 < len(val) <= 40:
                                            codes.append(val.strip())
                    except: pass
            except: pass
            return codes

        def extract_from_meta(html: str) -> list:
            """Extract product codes from meta tags."""
            codes = []
            try:
                soup_m = BS(html, "html.parser")
                meta_names = ["product:sku","product:id","sku","og:sku","article:code",
                              "product_id","item_number","part_number","mpn","cikkszam"]
                for name in meta_names:
                    for m in soup_m.find_all("meta", attrs={"name": name}):
                        val = m.get("content","").strip()
                        if val and 2 < len(val) <= 40: codes.append(val)
                    for m in soup_m.find_all("meta", attrs={"property": name}):
                        val = m.get("content","").strip()
                        if val and 2 < len(val) <= 40: codes.append(val)
            except: pass
            return codes

        def extract_from_embedded_json(html: str) -> list:
            """Extract codes from embedded JavaScript objects / window.__data__ patterns."""
            import json as _json
            codes = []
            try:
                # Find JSON blobs in script tags
                patterns = [
                    _re.compile(r'"code"\s*:\s*"([A-Z0-9][A-Z0-9_\-./]{2,30})"'),
                    _re.compile(r'"sku"\s*:\s*"([A-Z0-9][A-Z0-9_\-./]{2,30})"'),
                    _re.compile(r'"mpn"\s*:\s*"([A-Z0-9][A-Z0-9_\-./]{2,30})"'),
                    _re.compile(r'"productCode"\s*:\s*"([A-Z0-9][A-Z0-9_\-./]{2,30})"'),
                    _re.compile(r'"articleNumber"\s*:\s*"([A-Z0-9][A-Z0-9_\-./]{2,30})"'),
                    _re.compile(r'"itemCode"\s*:\s*"([A-Z0-9][A-Z0-9_\-./]{2,30})"'),
                    _re.compile(r'"cikkszam"\s*:\s*"([^"]{2,40})"', _re.I),
                    # SAP Hybris/Commerce Cloud patterns
                    _re.compile(r'data-product-code=(["\']?)([A-Z0-9][A-Z0-9_\-.]{2,30})\\1'),
                    _re.compile(r'data-code=(["\']?)([A-Z0-9][A-Z0-9_\-.]{2,30})\\1'),
                ]
                for pat in patterns:
                    codes.extend(pat.findall(html))
            except: pass
            return [c for c in codes if 2 < len(c) <= 40]

        def get_next_page_url(base_url: str, next_page_idx: int, html: str) -> str:
            """
            Build next page URL.
            Convention (Intercars & most sites):
              - First page  = base URL with NO page param
              - Second page = page=1
              - Third page  = page=2
              So next_page_idx is already the correct value for the page= param.
            """
            from urllib.parse import urlparse, parse_qs, urlencode, urlunparse, urljoin

            # First try HTML next-link (most reliable)
            try:
                soup_p = BS(html, "html.parser")
                for sel in ['a[rel="next"]', 'a.next', 'a.pagination-next',
                            'li.next a', 'li.pagination__item--next a',
                            '.pagination a.active + a', 'a.page-link[aria-label*="ext"]']:
                    el = soup_p.select_one(sel)
                    if el and el.get('href') and el.get('href') not in ('#',''):
                        return urljoin(base_url, el['href'])
                nl = soup_p.find('a', attrs={'aria-label': _re.compile(r'next|következő|weiter|nasl', _re.I)})
                if nl and nl.get('href') and nl.get('href') != '#':
                    return urljoin(base_url, nl['href'])
            except: pass

            # Build from base URL
            parsed = urlparse(base_url)
            # Strip existing page/currentPage from base
            params = parse_qs(parsed.query, keep_blank_values=True)
            params.pop('page', None)
            params.pop('currentPage', None)

            # Add page param (page=1 = second page, page=2 = third, etc.)
            params['page'] = [str(next_page_idx)]
            new_query = urlencode({k: v[0] for k,v in params.items()})
            return urlunparse(parsed._replace(query=new_query))

        AUTO_SELECTORS = [
            '[data-sku]','[data-product-id]','[data-article]','[data-code]',
            '[data-item-number]','[data-part-number]','[data-cikkszam]',
            '.sku','.product-code','.article-code','.item-code','.cikkszam',
            '.product-id','.art-nr','.item-nr','.part-number',
            'span.sku','td.sku','div.sku',
            '[itemprop="sku"]','[itemprop="productID"]',
            # Table column detection (e.g. admin panels)
            'td:nth-child(2)','td.col-cikkszam','td.col-sku',
        ]

        # Excel/CSV export link patterns
        EXPORT_PATTERNS = [
            _re.compile(r'letölt', _re.I),
            _re.compile(r'export', _re.I),
            _re.compile(r'excel', _re.I),
            _re.compile(r'download', _re.I),
            _re.compile(r'\.xlsx?$', _re.I),
            _re.compile(r'\.csv$', _re.I),
        ]

        def parse_excel_bytes(data: bytes) -> list:
            """Extract codes from Excel file bytes."""
            try:
                from openpyxl import load_workbook
                wb = load_workbook(_io.BytesIO(data), read_only=True, data_only=True)
                codes = []
                for ws in wb.worksheets:
                    header = None
                    sku_col = None
                    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                        if row_idx == 0:
                            # Detect SKU column from header
                            header = [str(c or "").lower() for c in row]
                            for i, h in enumerate(header):
                                if any(k in h for k in ['cikkszám','cikkszam','sku','code',
                                                         'article','termék','product','item']):
                                    sku_col = i
                                    break
                            if sku_col is None and len(header) > 0:
                                sku_col = 0  # fallback: first column
                            continue
                        if sku_col is not None and sku_col < len(row):
                            val = str(row[sku_col] or "").strip()
                            if val and val.lower() not in ('none','nan','') and 2 < len(val) <= 40:
                                codes.append(val)
                wb.close()
                return codes
            except Exception as e:
                self._js(f"onExtractProgress('Excel parse hiba: {str(e)[:60]}', 0)")
                return []

        def parse_csv_bytes(data: bytes) -> list:
            """Extract codes from CSV file bytes."""
            import csv as _csv
            codes = []
            try:
                text = data.decode('utf-8-sig', errors='replace')
                reader = _csv.reader(_io.StringIO(text))
                header = None; sku_col = 0
                for row_idx, row in enumerate(reader):
                    if row_idx == 0:
                        header = [c.lower() for c in row]
                        for i, h in enumerate(header):
                            if any(k in h for k in ['cikkszám','cikkszam','sku','code','article']):
                                sku_col = i; break
                        continue
                    if row and sku_col < len(row):
                        val = row[sku_col].strip()
                        if val and 2 < len(val) <= 40:
                            codes.append(val)
            except: pass
            return codes

        with requests.Session() as session:
            session.headers.update({"User-Agent": "Mozilla/5.0"})

            # Login to source site if needed
            if user and password:
                for src_url in urls:
                    try:
                        base = src_url.rstrip("/").rsplit("/", 1)[0] if "/" in src_url.replace("https://","").replace("http://","") else src_url
                        for login_path in ["/login", "/admin/login", "/bejelentkezes"]:
                            try:
                                lr = session.get(base + login_path, timeout=10)
                                if lr.status_code != 200: continue
                                soup = BS(lr.text, "html.parser")
                                form = soup.find("form")
                                if not form: continue
                                action = form.get("action") or (base + login_path)
                                if not action.startswith("http"): action = base + action
                                data = {inp.get("name",""):inp.get("value","")
                                        for inp in form.find_all("input") if inp.get("name")}
                                data["user-name"] = user; data["password"] = password
                                session.post(action, data=data, timeout=15, allow_redirects=True)
                                self._js("onExtractProgress('Bejelentkezve a forrás oldalra', 10)")
                                break
                            except: continue
                    except: pass

            for base_url in urls:
                self._js(f"onExtractProgress('Oldal elemzése: {base_url}', 15)")

                # ── STRATEGY 1: Detect Excel/CSV export link ────────────────
                try:
                    r = session.get(base_url, timeout=20)
                    if r.status_code == 200:
                        soup = BS(r.text, "html.parser")
                        export_url = None

                        # Look for download links/buttons
                        for el in soup.find_all(['a', 'button', 'input']):
                            href = el.get('href','') or el.get('action','') or ''
                            text = (el.get_text() or el.get('value','') or
                                    el.get('title','') or '').strip()
                            # Check if any export pattern matches href or text
                            combined = href + ' ' + text
                            if any(p.search(combined) for p in EXPORT_PATTERNS):
                                if href and href != '#':
                                    if not href.startswith('http'):
                                        from urllib.parse import urljoin
                                        href = urljoin(base_url, href)
                                    export_url = href
                                    self._js(f"onExtractProgress('Excel/CSV export link megtalálva: {text[:40]}', 20)")
                                    break

                        if export_url:
                            self._js(f"onExtractProgress('Letöltés: {export_url[:60]}', 25)")
                            er = session.get(export_url, timeout=60)
                            if er.status_code == 200:
                                ct = er.headers.get('content-type','').lower()
                                if 'excel' in ct or 'spreadsheet' in ct or 'xlsx' in export_url.lower() or 'xls' in export_url.lower():
                                    codes = parse_excel_bytes(er.content)
                                elif 'csv' in ct or 'csv' in export_url.lower():
                                    codes = parse_csv_bytes(er.content)
                                else:
                                    # Try Excel first, then CSV
                                    codes = parse_excel_bytes(er.content) or parse_csv_bytes(er.content)
                                if codes:
                                    self._js(f"onExtractProgress('Excel/CSV feldolgozva: {len(codes)} cikkszám', 90)")
                                    all_codes.extend(codes)
                                    total_pages += 1
                                    continue  # Skip page-by-page for this URL
                except Exception as e:
                    self._js(f"onExtractProgress('Export keresési hiba: {str(e)[:50]}', 15)")

                # ── STRATEGY 2-4: Page-by-page scraping ────────────────────
                seen_on_url = set()
                # Strip any existing page param from base URL — first page has none
                from urllib.parse import urlparse, parse_qs, urlencode, urlunparse
                _p = urlparse(base_url)
                _q = parse_qs(_p.query, keep_blank_values=True)
                _q.pop('page', None); _q.pop('currentPage', None)
                _clean_base = urlunparse(_p._replace(query=urlencode({k:v[0] for k,v in _q.items()})))

                page_url = _clean_base  # first page = no page param
                for page_num in range(1, max_pages + 1):
                    pass  # page_url updated in loop

                    try:
                        pct = 20 + int(page_num / max_pages * 70)
                        self._js(f"onExtractProgress('{page_num}. oldal: {page_url[:60]}', {pct})")
                        r = session.get(page_url, timeout=25,
                            headers={"Accept": "text/html,application/xhtml+xml",
                                     "Accept-Language": "hu-HU,hu;q=0.9,en;q=0.8"})
                        if r.status_code != 200: break
                        html = r.text
                        soup = BS(html, "html.parser")
                        page_codes = []

                        # Strategy 0: JSON-LD (schema.org) — highest precision
                        jld_codes = extract_from_jsonld(html)
                        if jld_codes:
                            page_codes.extend(jld_codes)

                        # Strategy 0b: Embedded JSON / data attributes
                        if not page_codes:
                            page_codes.extend(extract_from_embedded_json(html))

                        # Strategy 0c: Meta tags
                        if not page_codes:
                            page_codes.extend(extract_from_meta(html))

                        # Strategy 1: user-provided selector
                        if not page_codes and selector:
                            try:
                                for el in soup.select(selector):
                                    val = (el.get("data-sku") or el.get("data-code") or
                                           el.get("data-product-id") or el.get_text(strip=True))
                                    if val and 2 < len(val.strip()) <= 40:
                                        page_codes.append(val.strip())
                            except: pass

                        # Strategy 2: auto selectors
                        if not page_codes:
                            for sel_try in AUTO_SELECTORS:
                                try:
                                    for el in soup.select(sel_try):
                                        for attr in ["data-sku","data-code","data-product-id",
                                                     "data-product-code","data-cikkszam","data-articleno"]:
                                            val = el.get(attr,"").strip()
                                            if val and 2 < len(val) <= 40:
                                                page_codes.append(val)
                                        txt = el.get_text(strip=True)
                                        if txt and 2 < len(txt) <= 40 and CODE_PATTERN.match(txt):
                                            page_codes.append(txt)
                                    if page_codes: break
                                except: pass

                        # Strategy 3: table column detection (admin panels)
                        if not page_codes:
                            for tbl in soup.find_all('table'):
                                headers = [th.get_text(strip=True).lower()
                                          for th in tbl.find_all('th')]
                                sku_idx = next((i for i,h in enumerate(headers)
                                               if any(k in h for k in
                                                      ['cikkszám','cikkszam','sku','code','article','mpn'])), None)
                                if sku_idx is not None:
                                    for tr in tbl.find_all('tr')[1:]:
                                        tds = tr.find_all('td')
                                        if sku_idx < len(tds):
                                            val = tds[sku_idx].get_text(strip=True)
                                            if val and 2 < len(val) <= 40:
                                                page_codes.append(val)

                        # Strategy 4: automotive part regex
                        if not page_codes:
                            text = soup.get_text(" ")
                            for pat in AUTO_PART_PATTERNS:
                                matches = pat.findall(text)
                                if matches:
                                    page_codes.extend(matches[:500])
                                    break
                        # Strategy 5: general regex fallback
                        if not page_codes:
                            text = soup.get_text(" ")
                            page_codes = [m for m in CODE_PATTERN.findall(text)
                                         if not m.isdigit() or len(m) >= 6][:300]

                        if not page_codes: break
                        new_codes = [c for c in page_codes if c not in seen_on_url]
                        if not new_codes: break
                        seen_on_url.update(new_codes)
                        all_codes.extend(new_codes)
                        total_pages += 1

                        # Next page URL — page_num=1 → page=1, page_num=2 → page=2
                        if page_num < max_pages:
                            next_url = get_next_page_url(base_url, page_num, html)
                            if next_url and next_url != page_url:
                                page_url = next_url
                            else:
                                break

                    except Exception as e:
                        self._js(f"onExtractProgress('Oldal hiba: {str(e)[:50]}', 0)")
                        break

            # Deduplicate
            seen = set()
            unique = [c for c in all_codes if not (c in seen or seen.add(c))]
            self._js(f"onExtractProgress('Kész — {len(unique)} cikkszám', 100)")
            return {"codes": unique, "pages": total_pages}

        return {"codes": [], "pages": 0}

        all_codes = []
        total_pages = 0

        # Common patterns for product codes: uppercase+digits, with optional separators
        CODE_PATTERN = _re.compile(
            r'\b([A-Z][A-Z0-9]{1,6}[-_]?[A-Z0-9]{2,12}|[0-9]{4,14})\b'
        )
        # Common CSS selectors to try for product codes
        AUTO_SELECTORS = [
            '[data-sku]', '[data-product-id]', '[data-article]', '[data-code]',
            '[data-item-number]', '[data-part-number]',
            '.sku', '.product-code', '.article-code', '.item-code',
            '.product-id', '.part-number', '.art-nr', '.item-nr',
            'span.sku', 'td.sku', 'div.sku',
            '[itemprop="sku"]', '[itemprop="productID"]',
        ]

        with requests.Session() as session:
            session.headers.update({"User-Agent": "Mozilla/5.0"})

            # Login if needed
            if user and password:
                for url in urls:
                    try:
                        login_url = url.rstrip("/") + "/login"
                        r = session.get(login_url, timeout=15)
                        from bs4 import BeautifulSoup as _BS2
                        soup = _BS2(r.text, "html.parser")
                        form = soup.find("form")
                        if form:
                            action = form.get("action", login_url)
                            data = {inp.get("name",""): inp.get("value","")
                                    for inp in form.find_all("input") if inp.get("name")}
                            data["user-name"] = user
                            data["password"] = password
                            session.post(action, data=data, timeout=15)
                    except: pass

            for base_url in urls:
                seen_on_page = set()
                for page_num in range(1, max_pages + 1):
                    # Build paginated URL
                    if page_num == 1:
                        page_url = base_url
                    else:
                        # Try common pagination patterns
                        if '?' in base_url:
                            page_url = f"{base_url}&page={page_num}"
                        else:
                            page_url = f"{base_url}?page={page_num}"

                    try:
                        self._js(f"onExtractProgress('Oldal lekérése: {page_url}', {int(page_num/max_pages*80)})")
                        r = session.get(page_url, timeout=20)
                        if r.status_code != 200:
                            break
                        html = r.text
                        soup = BS(html, "html.parser")
                        page_codes = []

                        # Try specific selector first
                        if selector:
                            try:
                                for el in soup.select(selector):
                                    txt = el.get("data-sku") or el.get("data-code") or                                           el.get("data-product-id") or el.get_text(strip=True)
                                    if txt and len(txt) <= 30 and len(txt) >= 3:
                                        page_codes.append(txt.strip())
                            except: pass

                        # Try auto selectors
                        if not page_codes:
                            for sel in AUTO_SELECTORS:
                                try:
                                    els = soup.select(sel)
                                    for el in els:
                                        # Check data attributes first
                                        for attr in ["data-sku","data-code","data-product-id",
                                                     "data-article","data-item-number"]:
                                            val = el.get(attr, "").strip()
                                            if val and 2 < len(val) <= 30:
                                                page_codes.append(val)
                                        # Then text content
                                        txt = el.get_text(strip=True)
                                        if txt and 2 < len(txt) <= 30 and CODE_PATTERN.match(txt):
                                            page_codes.append(txt)
                                    if page_codes:
                                        break
                                except: pass

                        # Fallback: regex on full page text
                        if not page_codes:
                            text = soup.get_text(" ")
                            matches = CODE_PATTERN.findall(text)
                            # Filter out common false positives
                            page_codes = [m for m in matches
                                         if not m.isdigit() or len(m) >= 6][:500]

                        if not page_codes:
                            break  # No codes found, stop pagination

                        # Deduplicate within this URL
                        new_codes = [c for c in page_codes if c not in seen_on_page]
                        if not new_codes:
                            break  # Same page content, stop

                        seen_on_page.update(new_codes)
                        all_codes.extend(new_codes)
                        total_pages += 1

                    except Exception as e:
                        self._js(f"onExtractProgress('Hiba: {str(e)[:60]}', 0)")
                        break

            # Deduplicate all codes while preserving order
            seen = set()
            unique = [c for c in all_codes if not (c in seen or seen.add(c))]

            self._js(f"onExtractProgress('Kész — {len(unique)} cikkszám kinyerve', 100)")
            return {"codes": unique, "pages": total_pages}

        return {"codes": [], "pages": 0}

    # ── History ──────────────────────────────────────────────────────────
    def _history_path(self):
        hist_dir = os.path.join(os.path.expanduser("~"), "AppData", "Local",
                                "CikkChecker") if sys.platform=="win32" else                    os.path.join(os.path.expanduser("~"), ".cikkchecker")
        os.makedirs(hist_dir, exist_ok=True)
        return os.path.join(hist_dir, "history.json")

    def get_history(self):
        try:
            p = self._history_path()
            if os.path.exists(p):
                with open(p, encoding="utf-8") as f:
                    return json.load(f)
        except: pass
        return []

    def save_history(self, entry):
        try:
            history = self.get_history()
            history.append(entry)
            history = history[-100:]  # keep last 100 runs
            with open(self._history_path(), "w", encoding="utf-8") as f:
                json.dump(history, f, ensure_ascii=False, indent=2)
        except Exception as e:
            self._js(f"onLog({json.dumps(f'[WARN] Előzmény mentés hiba: {e}')})")
        return True

    # ── Windows notification ──────────────────────────────────────────────
    def notify(self, message: str, subtitle: str = ""):
        if sys.platform != "win32":
            return True
        try:
            # Use PowerShell toast notification
            ps_script = f"""
Add-Type -AssemblyName System.Windows.Forms
[System.Reflection.Assembly]::LoadWithPartialName('System.Windows.Forms') | Out-Null
$notify = New-Object System.Windows.Forms.NotifyIcon
$notify.Icon = [System.Drawing.SystemIcons]::Information
$notify.BalloonTipIcon = [System.Windows.Forms.ToolTipIcon]::Info
$notify.BalloonTipText = '{message.replace("'", "")}'
$notify.BalloonTipTitle = 'CikkChecker'
$notify.Visible = $True
$notify.ShowBalloonTip(5000)
Start-Sleep -Milliseconds 5500
$notify.Dispose()
"""
            subprocess.Popen(
                ["powershell", "-NonInteractive", "-WindowStyle", "Hidden",
                 "-Command", ps_script],
                creationflags=0x08000000,
                close_fds=True
            )
        except Exception as e:
            pass  # Notification is optional, never crash for it
        return True

    # ── Scheduler config ──────────────────────────────────────────────────
    def save_scheduler(self, config):
        self._cfg["scheduler"] = config
        save_config_file(self._cfg)
        return True

    def get_scheduler(self):
        return self._cfg.get("scheduler", {"enabled": False, "time": "08:00", "days": [1,2,3,4,5]})

    # ── JS helper ─────────────────────────────────────────────────────────
    def _js(self, code):
        if self._window:
            try: self._window.evaluate_js(code)
            except: pass


# ══════════════════════════════════════════════════════════════════════════════
# Entry point
# ══════════════════════════════════════════════════════════════════════════════
def main():
    api = Api()

    # Fix unknown dialog: JS calls back into Python
    original_handle = api._handle_unknown
    def patched_handle(code):
        if api._unk_mode == "skip_all": return "skip"
        if api._unk_mode == "stop_all": return "stop"

        api._unk_result = None
        api._js(f"""
            onUnknown({json.dumps(code)}).then(function(r) {{
                pywebview.api.set_unknown_result(r.action, r.applyAll);
            }});
        """)
        for _ in range(300):
            time.sleep(0.1)
            if api._unk_result is not None: break
        result = api._unk_result or {"action":"stop","applyAll":False}
        if result.get("applyAll"):
            if result["action"]=="skip": api._unk_mode="skip_all"
            else:                        api._unk_mode="stop_all"
        return result["action"]

    api._handle_unknown = patched_handle

    window = webview.create_window(
        title="CikkChecker",
        url=UI_PATH,
        js_api=api,
        width=1440, height=900,
        min_size=(1100, 700),
        background_color="#0D0821",
    )
    api.set_window(window)

    webview.start(debug=False)


if __name__ == "__main__":
    main()